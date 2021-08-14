from datetime import date
from tkinter import *
import shutil
import subprocess

from openpyxl import load_workbook
import numpy_financial as npf
import datetime
import openpyxl
path='../Pago-Cuotas.xlsx'

import win32gui, win32con

The_program_to_hide = win32gui.GetForegroundWindow()
win32gui.ShowWindow(The_program_to_hide , win32con.SW_HIDE)

def main():
  window = Tk()
  window.title("Huertas Inmobiliaria")
  window.geometry('400x250')

  lbl = Label(window, text="TABLA DE AMORTIZACIÓN" , font=("Arial", 15),bg='SeaGreen1' )
  lbl.grid( column=0, row=0)


  lbl = Label(window, text="Importe" , font=("Arial", 15) )
  lbl.grid(column=0, row=1)
  importe = Entry(window,width=10 ,font=("Arial", 15))
  importe.grid(column=1, row=1)

  lbl = Label(window, text="Inicial" ,font=("Arial", 15))
  lbl.grid(column=0, row=2)
  inicial = Entry(window,width=10 ,font=("Arial", 15))
  inicial.grid(column=1, row=2)

  lbl = Label(window, text="Coutas" ,font=("Arial", 15))
  lbl.grid(column=0, row=3)
  couta = Entry(window,width=10 ,font=("Arial", 15))
  couta.grid(column=1, row=3)

  lbl = Label(window, text="Interes" ,font=("Arial", 15))
  lbl.grid(column=0, row=4)
  Interes = Entry(window,width=10 ,font=("Arial", 15))
  Interes.grid(column=1, row=4)


  def selec():
    monitor.config(text = "Opción {}".format(opcion.get() ) )

  def clicked():
    capital = int(importe.get())
    tasa = int(Interes.get())
    ini =int(inicial.get())
    if(tasa>=1):
      tasa =int(Interes.get())/100
    plazo = int(couta.get())
    if(opcion.get()==1):
      fuente = "Plantilla/plantillaDolar.xlsx"
      
    if(opcion.get()==2):
      fuente = "Plantilla/plantillaSol.xlsx"
    resultado = "Pago-Cuotas.xlsx"
    
    shutil.copyfile(fuente, resultado)  
    Ejecutar(capital,ini,plazo,tasa)             

    commando=".\Pago-Cuotas.xlsx"
    subprocess.run(commando, shell=True)
      
  opcion = IntVar()

  lbl =Radiobutton(window, text="DOLARES", variable=opcion,value=1, command=selec,font=("Arial", 10))
  lbl.grid(column=0, row=5)
  lbl =Radiobutton(window, text="SOLES", variable=opcion,value=2, command=selec,font=("Arial", 10))
  lbl.grid(column=1, row=5)


  monitor = Label(window)

  btn = Button(window, text="Calcular",relief="groove", borderwidth=5, command=clicked , font=("Arial", 15) ,anchor="nw",bg='SeaGreen1' )

  btn.grid(column=1, row=6)



  window.mainloop()
def Ejecutar(importe,inicial,plazo,interes):
    capital=importe
    capital-=inicial
    cuota = round(npf.pmt(interes, plazo, -capital, 0), 0)
    datos = []
    saldo = capital
    today = date.today()
    prox=today
    for i in range(1, plazo+1):
        pago_capital = npf.ppmt(interes, i, plazo, -capital, 0)
        pago_int = cuota - pago_capital
        saldo -= pago_capital
        prox = prox + datetime.timedelta(days=30)  
        linea = [prox, format(cuota, '0,.0f'), format(pago_capital, '0,.0f'), format(pago_int, '0,.0f'), format(saldo, '0,.0f')]
        datos.append(linea)
    wb = load_workbook('Pago-Cuotas.xlsx')
    ws = wb['PAGO-CUOTAS']
    img = openpyxl.drawing.image.Image('logoCompany2.jpg')
    img.anchor = 'A1'
    ws.add_image(img)
    ws['B7'] = importe
    ws['B8'] = inicial
    ws['E7'] = plazo
    ws['E8'] = interes
    ws['B10'] = today
    ws['E10'] = today + datetime.timedelta(days=30)
    ws['E11'] = prox
    num=14
    lai=0
    for row in datos:
        num+=1
        lai+=1
        capo1=row[1].replace(',', '')
        capo2=row[2].replace(',', '')
        capo3=row[3].replace(',', '')
        capo4=row[4].replace(',', '')
        ws.cell(row=num, column=1, value=lai)
        ws.cell(row=num, column=2, value=row[0] )
        ws.cell(row=num, column=3, value=float(capo1))
        ws.cell(row=num, column=4, value=float(capo2))
        ws.cell(row=num, column=5, value=float(capo3))
        ws.cell(row=num, column=6, value=float(capo4))
    wb.save('Pago-Cuotas.xlsx')
if __name__ == '__main__':
    main()


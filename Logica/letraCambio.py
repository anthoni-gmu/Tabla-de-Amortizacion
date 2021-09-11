from openpyxl import load_workbook
from Logica.numeLetras import * 


from datetime import datetime
from dateutil.relativedelta import relativedelta
def letraCambio(fullname,dni,tabla,datos,moneda,domicilio,telefono,fiador,avalNombre,avalDomicilio,avalDOI,avalTelefono,banco,cuentabanco,cci):
    couta=tabla[2]
    fechas=[]
    precioCuota=0
    for aa in datos:
        fechas.append(aa[0].strftime('%d-%m-%Y'))
        precioCuota=aa[1].replace(',', '')
        precioCuota=float(precioCuota)  
            
    print(precioCuota)    
        
    wb = load_workbook('letras.xlsx')
    ws = wb['Hoja1']
     
    cuotainico=8
    fechainico=9
    
    
    giradorinic=16
    domicilio1inic=17
    doi1inic=18
    fiadorinic=19
    domicilio2inic=21
    doi2inic=22
    celda25=25
    num13=13
    
    if(moneda=="Dolar"):
        mei="$"
        meitext="Dolares Americanos"
    else:
        meitext="Soles"
        mei="S/"
        
    transforNum=numero_to_letras(int(precioCuota))
    print(transforNum)
    for xx in range(couta):
        
        celdacouta=f"D{cuotainico}"
        celdafecha=f"J{fechainico}"
        
        celdaContrato=f"I{fechainico}"
        
        celdagiradorinic=f"E{giradorinic}"
        celdadomicilio1inic=f"E{domicilio1inic}"
        celdadoi1inic=f"E{doi1inic}"
        celdafiadorinico=f"E{fiadorinic}"
        celdadomicilio2inic=f"E{domicilio2inic}"
        celdadoi2inic=f"E{doi2inic}"
        
        celdatel1inic=f"G{doi1inic}"
        celdatel2inic=f"G{doi2inic}"
        
        celdabanco=f"H{doi1inic}"
        celdacuenta=f"I{doi1inic}"
        celdacci=f"K{doi1inic}"
        
        celdarepre=f"F{celda25}"
        
        celdaimporte=f"K{cuotainico}"
        celdaimporteLetras=f"D{num13}"
        
        ws[celdaimporte]=f'{mei} {precioCuota}'
        ws[celdaimporteLetras]=f'{transforNum} {meitext}'
        fecc=aa[5].strftime('%d-%m-%Y')
        ws[celdafecha]=fechas[xx]
        ws[celdaContrato]=fecc

        
        ws[celdagiradorinic]=fullname
        ws[celdadomicilio1inic]=domicilio
        ws[celdadoi1inic]=dni
        ws[celdafiadorinico]=fiador
        ws[celdadomicilio2inic]=avalDomicilio
        ws[celdadoi2inic]=avalDOI
        ws[celdarepre]=avalNombre
        
        
        ws[celdatel1inic]=telefono
        ws[celdatel2inic]=avalTelefono
        
        ws[celdabanco]=banco
        ws[celdacuenta]=cuentabanco
        ws[celdacci]=cci
        
        
        
        ws[celdacouta]=f'{xx+1} - {couta}'
        fechainico+=26
        cuotainico+=26
        
        giradorinic+=26
        domicilio1inic+=26
        doi1inic+=26
        fiadorinic+=26
        domicilio2inic+=26
        doi2inic+=26
        num13+=26
        
        
       
   
    wb.save('letras.xlsx')
    

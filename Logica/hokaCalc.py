from openpyxl import load_workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta
def dataExcel(proyecto,fullname,dni,tabla,datos,fechaContrato,moneda,pendiente,lote):
    wb = load_workbook('Pago-Cuotas.xlsx')
    ws = wb['PAGO-CUOTAS']
    ws['B7'] = proyecto
    ws['B9'] = fullname
    ws['B10'] = dni
    ws['B12'] = tabla[0]
    ws['B13'] = tabla[1]
    ws['E12'] = pendiente
    ws['E13'] = tabla[2]
    ws['B16'] = fechaContrato
    ws['E16'] = tabla[3]
    ws['E7'] = lote
    
    nombrefrank=fullname.split()
    full=""
    for i in nombrefrank:
        nam=i[0]
        full+=nam
    
    init=0
    end=3
    proyecto =proyecto[init:end]
        
    cod=f'HI-{proyecto}-{full}'
    cod.upper()
    num=20
    lai=0
    for row in datos:
        num+=1
        lai+=1
        okay=f'{cod}-{lai}'
        capo1=row[1].replace(',', '')
        capo2=row[2].replace(',', '')
        interes=row[3].replace(',', '')
        capo4=row[4].replace(',', '')
        
        ws.cell(row=num, column=1, value=lai)
        ws.cell(row=num, column=2, value=okay.upper() )
        ws.cell(row=num, column=3, value=row[0].strftime('%Y-%m-%d'))
        ws.cell(row=num, column=4, value=float(capo1))
        ws.cell(row=num, column=5, value=float(capo4))
        
        
        ulti=row[0].strftime('%Y-%m-%d')
    ws['E17'] = ulti
    
    wb.save('Pago-Cuotas.xlsx')
    
